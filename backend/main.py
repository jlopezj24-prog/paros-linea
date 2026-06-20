"""API Paros de Línea — registro hora por hora, reporte gerencial y KPIs.

Reglas de negocio:
- Meta JPH = 62 (configurable por registro).
- Turnos: día (06:00-18:00) y noche (18:00-06:00). Lun-Sáb.
- 12 horas productivas - 45 min descanso (30 comedor + 15 snack) = 11h 15min reales.
- Categorías de paro: Operaciones (rojo), Mantenimiento (azul), Materiales (naranja), Programados (gris).
- Reporte gerencial: solo paros > 2 minutos no excluidos.
"""
from fastapi import FastAPI, Depends, HTTPException, Query, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from pathlib import Path
from sqlalchemy.orm import Session
from sqlalchemy import func
from pydantic import BaseModel, Field
from typing import Optional, List
from datetime import date as date_type, datetime
import io

from database import engine, get_db, Base, SessionLocal, IS_POSTGRES, DB_SCHEMA
import models
import seed
from sqlalchemy import text
import re

Base.metadata.create_all(bind=engine)

# Migración rápida: ampliar hora_label si quedó como VARCHAR(20) de un deploy viejo
if IS_POSTGRES:
    with engine.begin() as _c:
        _c.execute(text(
            f'ALTER TABLE "{DB_SCHEMA}".registros_hora '
            'ALTER COLUMN hora_label TYPE VARCHAR(80)'
        ))

app = FastAPI(title="Paros de Línea digital", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Contraseña simple para vista gerencial (cambiar via env GERENTE_PASS)
import os
GERENTE_PASS = os.environ.get("GERENTE_PASS", "gerente123")

HORAS_TURNO = {
    # (hora, label, minutos_productivos, tipo, meta_override_opcional)
    # tipo: "prod" = bloque productivo capturable; "break" = descanso, no capturable
    # Si meta_override está presente, sustituye al cálculo meta_jph * (minutos/60)
    "dia": [
        (1, "06:00-07:00", 60, "prod"),
        (2, "07:00-08:00", 60, "prod"),
        (3, "08:00-09:00", 60, "prod"),
        (4, "09:00-10:00", 60, "prod"),
        (5, "10:00-11:00", 60, "prod"),
        (0, "11:00-11:30  COMEDOR", 0, "break"),
        (6, "11:30-12:30", 60, "prod"),
        (7, "12:30-13:30", 60, "prod"),
        (8, "13:30-14:30", 60, "prod"),
        (9, "14:30-15:45  (incl. snack 15:00-15:15)", 60, "prod"),
        (10, "15:45-16:45", 60, "prod"),
        (11, "16:45-18:00", 75, "prod", 95),
    ],
    "noche": [
        (1, "18:00-19:00", 60, "prod"),
        (2, "19:00-20:00", 60, "prod"),
        (3, "20:00-21:00", 60, "prod"),
        (4, "21:00-22:00", 60, "prod"),
        (5, "22:00-23:00", 60, "prod"),
        (0, "23:00-23:30  COMEDOR", 0, "break"),
        (6, "23:30-00:30", 60, "prod"),
        (7, "00:30-01:30", 60, "prod"),
        (8, "01:30-02:30", 60, "prod"),
        (9, "02:30-03:45  (incl. snack 03:00-03:15)", 60, "prod"),
        (10, "03:45-04:45", 60, "prod"),
        (11, "04:45-06:00", 75, "prod", 95),
    ],
}


def _override_para(turno: str, hora: int):
    for item in HORAS_TURNO.get(turno, []):
        if item[0] == hora and item[3] == "prod":
            return item[4] if len(item) > 4 else None
    return None


@app.on_event("startup")
def _on_startup():
    seed.run()


# -------- Schemas --------
class LineaOut(BaseModel):
    id: int
    nombre: str
    area: str
    orden: int

    class Config:
        from_attributes = True


class CategoriaOut(BaseModel):
    id: int
    nombre: str
    color: str
    hex: str

    class Config:
        from_attributes = True


class ParoIn(BaseModel):
    categoria_id: int
    duracion_min: float = Field(gt=0)
    descripcion: str = ""


class ParoOut(BaseModel):
    id: int
    categoria_id: int
    categoria_nombre: str
    color: str
    hex: str
    duracion_min: float
    descripcion: str
    excluido_gerencial: bool

    class Config:
        from_attributes = True


class RegistroIn(BaseModel):
    fecha: date_type
    turno: str
    linea_id: int
    hora: int
    meta_jph: int = 62
    produccion: int = 0
    observaciones: str = ""
    paros: List[ParoIn] = []


class RegistroOut(BaseModel):
    id: int
    fecha: date_type
    turno: str
    linea_id: int
    linea_nombre: str
    area: str
    hora: int
    hora_label: str
    meta_jph: int
    meta: float  # meta efectiva del bloque (jobs)
    produccion: int
    minutos_disponibles: int
    observaciones: str
    paros: List[ParoOut]
    eficiencia: float  # produccion / meta_proporcional


def _meta_proporcional(meta_jph: int, minutos: int, override=None) -> float:
    if override is not None:
        return float(override)
    return round(meta_jph * (minutos / 60.0), 2)


def _meta_de(r) -> float:
    return _meta_proporcional(r.meta_jph, r.minutos_disponibles,
                              _override_para(r.turno, r.hora))


def _registro_to_out(r: models.RegistroHora) -> RegistroOut:
    meta_h = _meta_de(r)
    eficiencia = (r.produccion / meta_h * 100) if meta_h > 0 else 0
    return RegistroOut(
        id=r.id, fecha=r.fecha, turno=r.turno, linea_id=r.linea_id,
        linea_nombre=r.linea.nombre, area=r.linea.area, hora=r.hora,
        hora_label=r.hora_label, meta_jph=r.meta_jph, meta=meta_h,
        produccion=r.produccion,
        minutos_disponibles=r.minutos_disponibles, observaciones=r.observaciones or "",
        paros=[
            ParoOut(
                id=p.id, categoria_id=p.categoria_id,
                categoria_nombre=p.categoria.nombre, color=p.categoria.color,
                hex=p.categoria.hex, duracion_min=p.duracion_min,
                descripcion=p.descripcion or "",
                excluido_gerencial=p.excluido_gerencial,
            ) for p in r.paros
        ],
        eficiencia=round(eficiencia, 1),
    )


# -------- Catálogos --------
@app.get("/api/lineas", response_model=List[LineaOut])
def get_lineas(db: Session = Depends(get_db)):
    return db.query(models.Linea).filter_by(activa=True).order_by(models.Linea.orden).all()


@app.get("/api/categorias", response_model=List[CategoriaOut])
def get_categorias(db: Session = Depends(get_db)):
    return db.query(models.CategoriaParo).order_by(models.CategoriaParo.id).all()


@app.get("/api/horas/{turno}")
def get_horas(turno: str):
    if turno not in HORAS_TURNO:
        raise HTTPException(400, "Turno inválido (dia|noche)")
    out = []
    for item in HORAS_TURNO[turno]:
        h, l, m, t = item[0], item[1], item[2], item[3]
        override = item[4] if len(item) > 4 else None
        out.append({"hora": h, "label": l, "minutos": m, "tipo": t,
                    "meta_override": override})
    return out


@app.get("/api/turno-info")
def turno_info():
    return {
        "meta_jph_default": 62,
        "horas_brutas": 12,
        "min_descanso": 45,
        "horas_productivas": 11.25,
        "meta_turno": round(62 * 11.25, 0),
        "turnos": list(HORAS_TURNO.keys()),
    }


# -------- Registros --------
@app.post("/api/registros", response_model=RegistroOut)
def upsert_registro(payload: RegistroIn, db: Session = Depends(get_db)):
    if payload.turno not in HORAS_TURNO:
        raise HTTPException(400, "Turno inválido")
    horas_map = {
        item[0]: (item[1], item[2])
        for item in HORAS_TURNO[payload.turno] if item[3] == "prod"
    }
    if payload.hora not in horas_map:
        raise HTTPException(400, "Hora inválida (no es bloque productivo)")
    label, minutos = horas_map[payload.hora]

    r = (
        db.query(models.RegistroHora)
        .filter_by(fecha=payload.fecha, turno=payload.turno,
                   linea_id=payload.linea_id, hora=payload.hora)
        .first()
    )
    if not r:
        r = models.RegistroHora(
            fecha=payload.fecha, turno=payload.turno,
            linea_id=payload.linea_id, hora=payload.hora,
            hora_label=label, minutos_disponibles=minutos,
        )
        db.add(r)
    r.hora_label = label
    r.minutos_disponibles = minutos
    r.meta_jph = payload.meta_jph
    r.produccion = payload.produccion
    r.observaciones = payload.observaciones
    # Reemplazar paros
    for p in list(r.paros):
        db.delete(p)
    db.flush()
    for p in payload.paros:
        db.add(models.Paro(
            registro=r, categoria_id=p.categoria_id,
            duracion_min=p.duracion_min, descripcion=p.descripcion,
        ))
    db.commit()
    db.refresh(r)
    return _registro_to_out(r)


@app.get("/api/registros", response_model=List[RegistroOut])
def list_registros(
    fecha: date_type = Query(...),
    turno: str = Query(...),
    linea_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    q = db.query(models.RegistroHora).filter_by(fecha=fecha, turno=turno)
    if linea_id:
        q = q.filter_by(linea_id=linea_id)
    regs = q.order_by(models.RegistroHora.linea_id, models.RegistroHora.hora).all()
    return [_registro_to_out(r) for r in regs]


@app.delete("/api/registros/{registro_id}")
def delete_registro(registro_id: int, db: Session = Depends(get_db)):
    r = db.query(models.RegistroHora).get(registro_id)
    if not r:
        raise HTTPException(404)
    db.delete(r)
    db.commit()
    return {"ok": True}


# -------- Importación DTR (PDF Top Alarms) --------
def _hms_a_min(hms: str) -> float:
    """'00:01:05' -> minutos float (1.0833...)."""
    try:
        h, m, s = hms.strip().split(":")
        return round(int(h) * 60 + int(m) + int(s) / 60.0, 2)
    except Exception:
        return 0.0


def _detectar_categoria(alarm: str) -> str:
    """Devuelve nombre de categoría para una línea del DTR."""
    t = alarm.upper()
    # Llamada de operador / paros de operación en estación
    if "TEAM MEMBER HELP" in t:
        return "Operaciones"
    if " TT STOPPED" in t or t.startswith("TT STOPPED") \
       or " PP STOPPED" in t or t.startswith("PP STOPPED"):
        return "Operaciones"
    if "QF " in t or t.startswith("QF "):
        return "Operaciones"
    # Fallas mecánicas / proceso
    if t.startswith("MF ") or " MF " in t[:6] \
       or t.startswith("PF ") or t.startswith("TFIB ") \
       or t.startswith("TF "):
        return "Mantenimiento"
    return "Mantenimiento"


def _hora_bloque_desde_horas(turno: str, start_hhmm: str):
    """Busca en HORAS_TURNO[turno] el bloque productivo cuyo label empiece con start_hhmm."""
    for item in HORAS_TURNO.get(turno, []):
        if item[3] != "prod":
            continue
        if item[1].startswith(start_hhmm + "-"):
            return {"hora": item[0], "label": item[1], "minutos": item[2]}
    return None


def _sugerir_linea(texto: str, lineas) -> Optional[int]:
    """Heurística: busca tokens VES1..VES4, TRIM1..TRIM4 etc para sugerir línea."""
    t = texto.upper()
    mapeo = {
        "VES1": "Vestiduras 1", "VES2": "Vestiduras 2",
        "VES3": "Vestiduras 3", "VES4": "Vestiduras 4",
        "TRIM1": "Vestiduras 1", "TRIM2": "Vestiduras 2",
        "TRIM3": "Vestiduras 3", "TRIM4": "Vestiduras 4",
        "CHASIS1": "Chasis 1", "CHASIS2": "Chasis 2", "CHASIS3": "Chasis 3",
        "FINAL": "Linea Final", "MOTORES": "Motores", "AGVS": "AGVS",
        "PUERTAS": "Puertas", "TOLDOS": "Toldos",
    }
    for token, nombre_linea in mapeo.items():
        if token in t:
            for l in lineas:
                if l.nombre.lower() == nombre_linea.lower():
                    return l.id
    return None


@app.post("/api/dtr/parse")
async def parse_dtr_pdf(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """Parsea un PDF de Top Alarms (DTR) y devuelve los paros listos para precargar.

    Respuesta:
    {
      meta: { fecha, start_time, end_time, sub_area, turno_sugerido,
              hora_sugerida, hora_label, linea_sugerida_id },
      paros: [ { categoria_id, categoria_nombre, descripcion, duracion_min,
                 count, resource } ]
    }
    """
    if not (file.filename or "").lower().endswith(".pdf"):
        raise HTTPException(400, "Debe ser un archivo PDF")
    try:
        import pdfplumber
    except ImportError:
        raise HTTPException(500, "pdfplumber no instalado en el servidor")

    contenido = await file.read()
    texto_paginas = []
    try:
        with pdfplumber.open(io.BytesIO(contenido)) as pdf:
            for page in pdf.pages:
                texto_paginas.append(page.extract_text() or "")
    except Exception as e:
        raise HTTPException(400, f"PDF inválido: {e}")
    texto = "\n".join(texto_paginas)
    if not texto.strip():
        raise HTTPException(400, "No se pudo extraer texto del PDF")

    # --- Parseo de encabezado ---
    fecha_match = re.search(r"From:\s*(\d{1,2}/\d{1,2}/\d{2,4})", texto)
    start_match = re.search(r"Start Time:\s*(\d{1,2}:\d{2})", texto)
    end_match = re.search(r"End Time:\s*(\d{1,2}:\d{2})", texto)
    sub_area_match = re.search(r"Sub Area:\s*(\S+)", texto)

    fecha_iso = None
    if fecha_match:
        try:
            d = datetime.strptime(fecha_match.group(1), "%m/%d/%Y").date()
            fecha_iso = d.isoformat()
        except ValueError:
            try:
                d = datetime.strptime(fecha_match.group(1), "%m/%d/%y").date()
                fecha_iso = d.isoformat()
            except ValueError:
                pass

    start_hhmm = start_match.group(1) if start_match else None
    end_hhmm = end_match.group(1) if end_match else None
    sub_area = sub_area_match.group(1) if sub_area_match else ""

    # Turno: si Start está entre 06:00-17:59 → día; si no → noche
    turno_sug = None
    hora_info = None
    if start_hhmm:
        try:
            hh = int(start_hhmm.split(":")[0])
            turno_sug = "dia" if 6 <= hh < 18 else "noche"
            hora_info = _hora_bloque_desde_horas(turno_sug, start_hhmm)
        except ValueError:
            pass

    # Línea sugerida
    lineas = db.query(models.Linea).filter_by(activa=True).all()
    linea_sug_id = _sugerir_linea(texto, lineas)

    # Categorías por nombre → id
    categorias = db.query(models.CategoriaParo).all()
    cat_por_nombre = {c.nombre: c for c in categorias}

    # --- Parseo de filas de alarmas ---
    # Formato típico:
    # "G1TRIM4C_VES4 MF GT_SKL04 IFD404 IDC2 EN FALLA @A15N REVISAR ARMORSTART 1 00:01:05 00:01:05"
    # → resource | alarm_message | count | high_alarm | total_adjusted_duration
    fila_re = re.compile(
        r"^(?P<res>\S+)\s+"
        r"(?P<msg>.+?)\s+"
        r"(?P<count>\d+)\s+"
        r"(?P<high>\d{1,2}:\d{2}:\d{2})\s+"
        r"(?P<total>\d{1,2}:\d{2}:\d{2})\s*$"
    )
    paros_out = []
    for raw_line in texto.split("\n"):
        line = raw_line.strip()
        if not line:
            continue
        m = fila_re.match(line)
        if not m:
            continue
        # Filtrar líneas de totales o headers que pudieran matchear casualmente
        msg = m.group("msg").strip()
        if msg.lower().startswith("alarm message"):
            continue
        if "highest alarm" in line.lower() or "report generated" in line.lower():
            continue
        duracion = _hms_a_min(m.group("total"))
        if duracion <= 1.0:
            continue
        cat_nombre = _detectar_categoria(msg)
        cat = cat_por_nombre.get(cat_nombre) or cat_por_nombre.get("Mantenimiento")
        if cat is None:
            continue
        paros_out.append({
            "categoria_id": cat.id,
            "categoria_nombre": cat.nombre,
            "descripcion": msg[:300],
            "duracion_min": duracion,
            "count": int(m.group("count")),
            "resource": m.group("res"),
        })

    return {
        "meta": {
            "fecha": fecha_iso,
            "start_time": start_hhmm,
            "end_time": end_hhmm,
            "sub_area": sub_area,
            "turno_sugerido": turno_sug,
            "hora_sugerida": hora_info["hora"] if hora_info else None,
            "hora_label": hora_info["label"] if hora_info else None,
            "minutos_bloque": hora_info["minutos"] if hora_info else None,
            "linea_sugerida_id": linea_sug_id,
        },
        "paros": paros_out,
    }


# -------- Dashboard DTR (acumulado) --------
# Categorías DTR según el PDF Top Alarms (NO confundir con CategoriaParo)
DTR_CATEGORIAS = ["FPS", "Andon", "PF", "MF", "TFS", "TFIB", "Otros"]
DTR_CATEGORIAS_DESC = {
    "FPS": "Error Proofing",
    "Andon": "Andon (Help Call)",
    "PF": "Over travel",
    "MF": "Mantenimiento",
    "TFS": "Vacíos (Starvation)",
    "TFIB": "Bloqueos (Blocking)",
    "Otros": "Otros",
}


def _clasificar_dtr(mensaje: str) -> str:
    """Clasifica una alarma DTR en una de las 6 categorías + Otros."""
    t = (mensaje or "").upper()
    if "STOPPED@FPS" in t:
        return "FPS"
    if "TEAM MEMBER HELP CALL FAULT" in t:
        return "Andon"
    # TFIB antes que TFS porque "TFIB" contiene "TF"
    if t.startswith("TFIB ") or " TFIB " in t:
        return "TFIB"
    if t.startswith("TFS ") or " TFS " in t:
        return "TFS"
    if t.startswith("PF ") or " PF " in t[:5]:
        return "PF"
    if t.startswith("MF ") or " MF " in t[:5]:
        return "MF"
    return "Otros"


def _parsear_pdf_dtr(contenido: bytes):
    """Devuelve (texto_completo, lista_de_filas) donde fila = dict con resource, mensaje, count, duracion_min."""
    try:
        import pdfplumber
    except ImportError:
        raise HTTPException(500, "pdfplumber no instalado en el servidor")
    texto_paginas = []
    try:
        with pdfplumber.open(io.BytesIO(contenido)) as pdf:
            for page in pdf.pages:
                texto_paginas.append(page.extract_text() or "")
    except Exception as e:
        raise HTTPException(400, f"PDF inválido: {e}")
    texto = "\n".join(texto_paginas)
    fila_re = re.compile(
        r"^(?P<res>\S+)\s+"
        r"(?P<msg>.+?)\s+"
        r"(?P<count>\d+)\s+"
        r"(?P<high>\d{1,2}:\d{2}:\d{2})\s+"
        r"(?P<total>\d{1,2}:\d{2}:\d{2})\s*$"
    )
    filas = []
    for raw in texto.split("\n"):
        line = raw.strip()
        if not line:
            continue
        m = fila_re.match(line)
        if not m:
            continue
        msg = m.group("msg").strip()
        if msg.lower().startswith("alarm message"):
            continue
        if "highest alarm" in line.lower() or "report generated" in line.lower():
            continue
        dur = _hms_a_min(m.group("total"))
        if dur <= 0:
            continue
        filas.append({
            "resource": m.group("res"),
            "mensaje": msg,
            "count": int(m.group("count")),
            "duracion_min": dur,
        })
    return texto, filas


def _meta_dtr_desde_texto(texto: str):
    """Extrae fecha, start_time, end_time, sub_area del encabezado del PDF."""
    fecha_match = re.search(r"From:\s*(\d{1,2}/\d{1,2}/\d{2,4})", texto)
    start_match = re.search(r"Start Time:\s*(\d{1,2}:\d{2})", texto)
    end_match = re.search(r"End Time:\s*(\d{1,2}:\d{2})", texto)
    sub_area_match = re.search(r"Sub Area:\s*(\S+)", texto)
    fecha_iso = None
    if fecha_match:
        for fmt in ("%m/%d/%Y", "%m/%d/%y"):
            try:
                fecha_iso = datetime.strptime(fecha_match.group(1), fmt).date().isoformat()
                break
            except ValueError:
                pass
    return {
        "fecha": fecha_iso,
        "start_time": start_match.group(1) if start_match else "",
        "end_time": end_match.group(1) if end_match else "",
        "sub_area": sub_area_match.group(1) if sub_area_match else "",
    }


@app.post("/api/dtr/ingest")
async def dtr_ingest(
    file: UploadFile = File(...),
    fecha: date_type = Query(...),
    turno: str = Query(...),  # dia | noche
    linea_id: int = Query(...),
    db: Session = Depends(get_db),
):
    """Sube un PDF DTR y lo persiste para el dashboard acumulado.

    Reemplaza cualquier import previo con la misma (fecha, turno, linea_id).
    """
    if turno not in ("dia", "noche"):
        raise HTTPException(400, "turno debe ser 'dia' o 'noche'")
    if not (file.filename or "").lower().endswith(".pdf"):
        raise HTTPException(400, "Debe ser un archivo PDF")
    linea = db.query(models.Linea).filter_by(id=linea_id).first()
    if not linea:
        raise HTTPException(404, "Línea no encontrada")

    contenido = await file.read()
    texto, filas = _parsear_pdf_dtr(contenido)
    if not filas:
        raise HTTPException(400, "No se detectaron alarmas en el PDF")

    meta_pdf = _meta_dtr_desde_texto(texto)

    # Reemplazar import previo si existe
    prev = (
        db.query(models.DTRImport)
        .filter_by(fecha=fecha, turno=turno, linea_id=linea_id)
        .first()
    )
    if prev:
        db.delete(prev)
        db.flush()

    imp = models.DTRImport(
        fecha=fecha, turno=turno, linea_id=linea_id,
        sub_area=meta_pdf["sub_area"][:40],
        start_time=meta_pdf["start_time"][:10],
        end_time=meta_pdf["end_time"][:10],
        archivo_nombre=(file.filename or "")[:200],
    )
    db.add(imp)
    db.flush()

    for f in filas:
        cat = _clasificar_dtr(f["mensaje"])
        db.add(models.DTRAlarma(
            import_id=imp.id,
            resource=f["resource"][:80],
            mensaje=f["mensaje"][:1000],
            categoria_dtr=cat,
            count=f["count"],
            duracion_min=f["duracion_min"],
        ))
    db.commit()
    db.refresh(imp)

    return {
        "id": imp.id,
        "fecha": imp.fecha.isoformat(),
        "turno": imp.turno,
        "linea_id": imp.linea_id,
        "linea_nombre": linea.nombre,
        "sub_area": imp.sub_area,
        "start_time": imp.start_time,
        "end_time": imp.end_time,
        "alarmas_total": len(filas),
        "duracion_total_min": round(sum(f["duracion_min"] for f in filas), 2),
    }


@app.get("/api/dtr/imports")
def dtr_imports_list(
    fecha_desde: Optional[date_type] = None,
    fecha_hasta: Optional[date_type] = None,
    linea_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    q = db.query(models.DTRImport)
    if fecha_desde:
        q = q.filter(models.DTRImport.fecha >= fecha_desde)
    if fecha_hasta:
        q = q.filter(models.DTRImport.fecha <= fecha_hasta)
    if linea_id:
        q = q.filter(models.DTRImport.linea_id == linea_id)
    imps = q.order_by(models.DTRImport.fecha.desc(),
                      models.DTRImport.turno).all()
    out = []
    for i in imps:
        total_min = sum(a.duracion_min for a in i.alarmas)
        out.append({
            "id": i.id,
            "fecha": i.fecha.isoformat(),
            "turno": i.turno,
            "linea_id": i.linea_id,
            "linea_nombre": i.linea.nombre,
            "sub_area": i.sub_area,
            "start_time": i.start_time,
            "end_time": i.end_time,
            "archivo_nombre": i.archivo_nombre,
            "alarmas_total": len(i.alarmas),
            "duracion_total_min": round(total_min, 2),
            "created_at": i.created_at.isoformat() if i.created_at else None,
        })
    return out


@app.delete("/api/dtr/imports/{imp_id}")
def dtr_import_delete(imp_id: int, db: Session = Depends(get_db)):
    imp = db.query(models.DTRImport).get(imp_id)
    if not imp:
        raise HTTPException(404)
    db.delete(imp)
    db.commit()
    return {"ok": True}


@app.get("/api/dtr/dashboard")
def dtr_dashboard(
    fecha_desde: Optional[date_type] = None,
    fecha_hasta: Optional[date_type] = None,
    linea_id: Optional[int] = None,
    categoria: Optional[str] = None,  # filtro por categoría DTR
    top_n: int = Query(15, ge=1, le=100),
    db: Session = Depends(get_db),
):
    """Métricas agregadas para el dashboard DTR.

    Devuelve:
      resumen: { total_paros, total_min, total_imports }
      por_categoria: [{categoria, descripcion, count_total, duracion_min}]
      top_alarmas_duracion: [{mensaje, count, duracion_min, categoria}]
      top_alarmas_frecuencia: [{...}]
      tendencia_diaria: [{fecha, duracion_min, count_total}]
    """
    q = (
        db.query(models.DTRAlarma, models.DTRImport)
        .join(models.DTRImport, models.DTRAlarma.import_id == models.DTRImport.id)
    )
    if fecha_desde:
        q = q.filter(models.DTRImport.fecha >= fecha_desde)
    if fecha_hasta:
        q = q.filter(models.DTRImport.fecha <= fecha_hasta)
    if linea_id:
        q = q.filter(models.DTRImport.linea_id == linea_id)
    if categoria and categoria in DTR_CATEGORIAS:
        q = q.filter(models.DTRAlarma.categoria_dtr == categoria)

    rows = q.all()
    if not rows:
        return {
            "resumen": {"total_paros": 0, "total_min": 0.0, "total_imports": 0},
            "por_categoria": [],
            "top_alarmas_duracion": [],
            "top_alarmas_frecuencia": [],
            "tendencia_diaria": [],
        }

    # Resumen global
    total_min = sum(a.duracion_min for a, _ in rows)
    total_paros = sum(a.count for a, _ in rows)
    total_imports = len({i.id for _, i in rows})

    # Por categoría
    cat_acum = {c: {"count_total": 0, "duracion_min": 0.0} for c in DTR_CATEGORIAS}
    for a, _ in rows:
        bucket = cat_acum.get(a.categoria_dtr) or cat_acum["Otros"]
        bucket["count_total"] += a.count
        bucket["duracion_min"] += a.duracion_min
    por_categoria = []
    for c in DTR_CATEGORIAS:
        v = cat_acum[c]
        por_categoria.append({
            "categoria": c,
            "descripcion": DTR_CATEGORIAS_DESC[c],
            "count_total": v["count_total"],
            "duracion_min": round(v["duracion_min"], 2),
        })

    # Top alarmas — agrupa por mensaje (normalizado a 200 chars)
    agrup = {}
    for a, _ in rows:
        key = (a.categoria_dtr, a.mensaje[:200])
        if key not in agrup:
            agrup[key] = {
                "mensaje": a.mensaje[:200],
                "categoria": a.categoria_dtr,
                "count": 0, "duracion_min": 0.0,
            }
        agrup[key]["count"] += a.count
        agrup[key]["duracion_min"] += a.duracion_min
    lista = list(agrup.values())
    for it in lista:
        it["duracion_min"] = round(it["duracion_min"], 2)
    top_dur = sorted(lista, key=lambda x: x["duracion_min"], reverse=True)[:top_n]
    top_freq = sorted(lista, key=lambda x: x["count"], reverse=True)[:top_n]

    # Tendencia diaria
    dia_acum = {}
    for a, i in rows:
        k = i.fecha.isoformat()
        if k not in dia_acum:
            dia_acum[k] = {"fecha": k, "duracion_min": 0.0, "count_total": 0}
        dia_acum[k]["duracion_min"] += a.duracion_min
        dia_acum[k]["count_total"] += a.count
    tendencia = sorted(dia_acum.values(), key=lambda x: x["fecha"])
    for d in tendencia:
        d["duracion_min"] = round(d["duracion_min"], 2)

    return {
        "resumen": {
            "total_paros": total_paros,
            "total_min": round(total_min, 2),
            "total_imports": total_imports,
        },
        "por_categoria": por_categoria,
        "top_alarmas_duracion": top_dur,
        "top_alarmas_frecuencia": top_freq,
        "tendencia_diaria": tendencia,
    }

# -------- Reporte gerencial --------
def _check_gerente(password: str):
    # Acceso libre: validación eliminada por solicitud del usuario.
    return


@app.get("/api/reporte-gerencial")
def reporte_gerencial(
    fecha: date_type = Query(...),
    turno: str = Query(...),
    password: str = Query(...),
    umbral_min: float = 2.0,
    linea_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    _check_gerente(password)
    q = (
        db.query(models.Paro, models.RegistroHora, models.Linea, models.CategoriaParo)
        .join(models.RegistroHora, models.Paro.registro_id == models.RegistroHora.id)
        .join(models.Linea, models.RegistroHora.linea_id == models.Linea.id)
        .join(models.CategoriaParo, models.Paro.categoria_id == models.CategoriaParo.id)
        .filter(models.RegistroHora.fecha == fecha)
        .filter(models.RegistroHora.turno == turno)
        .filter(models.Paro.duracion_min > umbral_min)
        .filter(models.Paro.excluido_gerencial == False)  # noqa: E712
    )
    if linea_id:
        q = q.filter(models.RegistroHora.linea_id == linea_id)
    rows = q.order_by(models.Linea.orden, models.RegistroHora.hora).all()
    return [
        {
            "paro_id": p.id,
            "linea": l.nombre,
            "area": l.area,
            "hora": r.hora_label,
            "categoria": c.nombre,
            "color": c.color,
            "hex": c.hex,
            "duracion_min": p.duracion_min,
            "descripcion": p.descripcion or "",
        }
        for p, r, l, c in rows
    ]


@app.patch("/api/paros/{paro_id}/excluir")
def excluir_paro(paro_id: int, password: str = Query(...), db: Session = Depends(get_db)):
    _check_gerente(password)
    p = db.query(models.Paro).get(paro_id)
    if not p:
        raise HTTPException(404)
    p.excluido_gerencial = True
    db.commit()
    return {"ok": True}


@app.patch("/api/paros/{paro_id}/restaurar")
def restaurar_paro(paro_id: int, password: str = Query(...), db: Session = Depends(get_db)):
    _check_gerente(password)
    p = db.query(models.Paro).get(paro_id)
    if not p:
        raise HTTPException(404)
    p.excluido_gerencial = False
    db.commit()
    return {"ok": True}


# -------- KPIs / Dashboard --------
@app.get("/api/kpis")
def kpis(fecha: date_type = Query(...), turno: str = Query(...), db: Session = Depends(get_db)):
    regs = db.query(models.RegistroHora).filter_by(fecha=fecha, turno=turno).all()
    total_prod = sum(r.produccion for r in regs)
    total_meta = sum(_meta_de(r) for r in regs)
    # Bloques productivos esperados por turno
    horas_esperadas = sum(
        1 for item in HORAS_TURNO.get(turno, []) if item[3] == "prod"
    )
    # Por línea: incluir TODAS las líneas activas para detectar faltantes
    lineas = db.query(models.Linea).filter_by(activa=True).order_by(models.Linea.orden).all()
    por_linea = {
        l.nombre: {
            "linea": l.nombre, "area": l.area,
            "produccion": 0, "meta": 0.0, "paros_min": 0.0,
            "horas_capturadas": 0, "horas_esperadas": horas_esperadas,
        }
        for l in lineas
    }
    for r in regs:
        d = por_linea.get(r.linea.nombre)
        if not d:
            continue
        d["produccion"] += r.produccion
        d["meta"] += _meta_de(r)
        d["paros_min"] += sum(p.duracion_min for p in r.paros)
        d["horas_capturadas"] += 1
    for d in por_linea.values():
        d["eficiencia"] = round(d["produccion"] / d["meta"] * 100, 1) if d["meta"] else 0
        d["meta"] = round(d["meta"], 1)
        d["pendientes"] = max(0, d["horas_esperadas"] - d["horas_capturadas"])
        d["completo"] = d["horas_capturadas"] >= d["horas_esperadas"]
    # Por categoría
    por_cat = {}
    for r in regs:
        for p in r.paros:
            k = p.categoria.nombre
            d = por_cat.setdefault(k, {"categoria": k, "color": p.categoria.color,
                                       "hex": p.categoria.hex, "minutos": 0, "eventos": 0})
            d["minutos"] += p.duracion_min
            d["eventos"] += 1
    return {
        "total_produccion": total_prod,
        "total_meta": round(total_meta, 1),
        "cumplimiento_pct": round(total_prod / total_meta * 100, 1) if total_meta else 0,
        "por_linea": list(por_linea.values()),
        "por_categoria": list(por_cat.values()),
    }


# -------- Export Excel --------
@app.get("/api/export/excel")
def export_excel(fecha: date_type = Query(...), turno: str = Query(...),
                 db: Session = Depends(get_db)):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    regs = (
        db.query(models.RegistroHora)
        .filter_by(fecha=fecha, turno=turno)
        .join(models.Linea).order_by(models.Linea.orden, models.RegistroHora.hora)
        .all()
    )
    wb = Workbook()
    ws = wb.active
    ws.title = f"{fecha}_{turno}"
    headers = ["Área", "Línea", "Hora", "Meta JPH", "Producción",
               "Eficiencia %", "Paros (categoría: min)", "Observaciones"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
    color_map = {"red": "FFDC2626", "blue": "FF2563EB",
                 "orange": "FFEA580C", "gray": "FF6B7280"}
    for r in regs:
        meta_h = _meta_de(r)
        ef = round(r.produccion / meta_h * 100, 1) if meta_h else 0
        paros_txt = " | ".join(
            f"{p.categoria.nombre}: {p.duracion_min}m" for p in r.paros
        )
        ws.append([r.linea.area, r.linea.nombre, r.hora_label, r.meta_jph,
                   r.produccion, ef, paros_txt, r.observaciones or ""])
        # color de fondo si hay paros (primer categoría)
        if r.paros:
            color = color_map.get(r.paros[0].categoria.color)
            if color:
                ws.cell(row=ws.max_row, column=7).fill = PatternFill(
                    fill_type="solid", fgColor=color)
                ws.cell(row=ws.max_row, column=7).font = Font(color="FFFFFFFF", bold=True)
    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = f"paros_{fecha}_{turno}.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@app.get("/api/export/gerencial-excel")
def export_gerencial_excel(
    fecha: date_type = Query(...),
    turno: str = Query(...),
    password: str = Query(...),
    umbral_min: float = 2.0,
    linea_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    _check_gerente(password)
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    q = (
        db.query(models.Paro, models.RegistroHora, models.Linea, models.CategoriaParo)
        .join(models.RegistroHora, models.Paro.registro_id == models.RegistroHora.id)
        .join(models.Linea, models.RegistroHora.linea_id == models.Linea.id)
        .join(models.CategoriaParo, models.Paro.categoria_id == models.CategoriaParo.id)
        .filter(models.RegistroHora.fecha == fecha)
        .filter(models.RegistroHora.turno == turno)
        .filter(models.Paro.duracion_min > umbral_min)
        .filter(models.Paro.excluido_gerencial == False)  # noqa: E712
    )
    if linea_id:
        q = q.filter(models.RegistroHora.linea_id == linea_id)
    rows = q.order_by(models.Linea.orden, models.RegistroHora.hora).all()
    wb = Workbook()
    ws = wb.active
    ws.title = f"Gerencial_{turno}"
    ws.append([f"Reporte Gerencial · {fecha} · Turno {turno}",
               f"Paros > {umbral_min} min"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.append([])
    headers = ["Área", "Línea", "Hora", "Categoría", "Min", "Descripción"]
    ws.append(headers)
    for c in ws[3]:
        c.font = Font(bold=True, color="FFFFFFFF")
        c.fill = PatternFill(fill_type="solid", fgColor="FF1E293B")
        c.alignment = Alignment(horizontal="center")
    color_map = {"red": "FFDC2626", "blue": "FF2563EB",
                 "orange": "FFEA580C", "gray": "FF6B7280"}
    total_min = 0
    for p, r, l, c in rows:
        ws.append([l.area, l.nombre, r.hora_label, c.nombre,
                   p.duracion_min, p.descripcion or ""])
        fill_color = color_map.get(c.color)
        if fill_color:
            cell = ws.cell(row=ws.max_row, column=4)
            cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)
            cell.font = Font(color="FFFFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        total_min += p.duracion_min
    ws.append([])
    ws.append(["", "", "", "TOTAL", round(total_min, 1), f"{len(rows)} paros"])
    ws.cell(row=ws.max_row, column=4).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=5).font = Font(bold=True)
    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    fname = f"reporte_gerencial_{fecha}_{turno}.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@app.get("/api")
def api_root():
    return {"app": "Paros de Línea digital", "docs": "/docs"}


# --- Servir frontend buildeado (modo producción) ---
# Si existe ../frontend/dist, montar como SPA estática en la raíz.
_DIST = Path(__file__).resolve().parent.parent / "frontend" / "dist"
if _DIST.exists():
    # Assets (JS/CSS) servidos con caché por Vite
    app.mount("/assets", StaticFiles(directory=_DIST / "assets"), name="assets")

    @app.get("/", include_in_schema=False)
    def spa_index():
        return FileResponse(_DIST / "index.html")

    @app.get("/{full_path:path}", include_in_schema=False)
    def spa_fallback(full_path: str):
        # Cualquier ruta no /api ni /docs ni /assets devuelve index.html (React Router)
        if full_path.startswith(("api/", "docs", "openapi.json")):
            raise HTTPException(status_code=404)
        target = _DIST / full_path
        if full_path and target.is_file():
            return FileResponse(target)
        return FileResponse(_DIST / "index.html")
else:
    @app.get("/")
    def root():
        return {"app": "Paros de Línea digital", "docs": "/docs", "warning": "frontend/dist no encontrado"}
